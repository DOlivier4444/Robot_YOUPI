from enum import Enum
import os
import openpyxl.worksheet
import openpyxl.worksheet.worksheet
import serial #type:ignore
import openpyxl #https://www.datacamp.com/tutorial/python-excel-tutorial  #type:ignore
import time
import math
import numpy as np #type:ignore


# --------------------- Classes ---------------------
## 
class Excel_Params :
  PROGRAM_FILE_NAME = 'programs.xlsx'
  STARTING_ROW = 4 # row where the program start in the xlsx file

## Communications
class Comms :
  class To_Arduino(Enum) :
    RASPBERRY_READY   = 0xA1
    START_MOVEMENT    = 0xA2
    ABORT_PROGRAM     = 0xA3
    PROGRAM_EXECUTION = 0xA4
    PROGRAM_FINISHED  = 0xA5
    ROBOT_RESET_SIGNAL= 0xAA
    RESET_ARDUINO     = 0xAB

  class From_Arduino(Enum) :
    ARDUINO_READY     = 0xB1
    READY_TO_RECEIVE  = 0xB2
    MOVEMENT_FINISHED = 0xB3

  #class General(Enum) :
  #  ERROR_MOVEMENT    = 0xFF

## Robot's parameters
L1 = 0.28
L2 = 0.162
L3 = 0.162
L4 = 0.15
class Youpi :
  class Motors(Enum) :
    M0_BASE     = 0  # motor0 - base
    M1_SHOULDER = 1  # motor1 - shoulder
    M2_ELBOW    = 2  # motor2 - elbow
    M3_PITCH    = 3  # motor3 - wrist pitch
    M4_ROLL     = 4  # motor4 - wrist roll
    M5_GRIPPER  = 5  # motor5 - gripper

  class Robot_Arms(Enum) :
    # Lengths of the Robot's arm
    L1 = L1
    L2 = L2
    L3 = L3
    L4 = L4

  DH_Params = {
    
    # DH parameters
    #
    # const float D1 = L1             ; const float A1 = 0.0            ; const float ALPHA1 = 90.0 ;
    # const float D2 = 0.0            ; const float A2 = L2             ; const float ALPHA2 = 0.0  ;
    # const float D3 = 0.0            ; const float A3 = L3             ; const float ALPHA3 = 0.0  ;
    # const float D4 = 0.0            ; const float A4 = 0.0            ; const float ALPHA4 = 90.0 ;
    # const float D5 = L4             ; const float A5 = 0.0            ; const float ALPHA5 = 0.0  ;
    # const float D6 = penOffsetV     ; const float A6 = penOffsetH     ; const float ALPHA6 = 0.0  ;
    "D"     : [L1   , 0.0 , 0.0 , 0.0  , L4  , 0.0], #[5] = penOffsetV
    "A"     : [0.0  , L2  , L3  , 0.0  , 0.0 , 0.0], #[5] = penOffsetH
    "ALPHA" : [90.0 , 0.0 , 0.0 , 90.0 , 0.0 , 0.0]
  }

  class movement_Mode(str, Enum) :
    JOINT = "J"
    LINEAR = "L"

# --------------------- Functions ---------------------
def Serial_Initialisation() -> serial :
  arduino = Serial_Connexion()
  Serial_Synchronisation(arduino)
  return arduino

## Serial communication
def Serial_Connexion() -> serial: # Test each ports of the RPI to connect to the Arduino
  ports = ["ACM0", "ACM1", "USB0", "USB1"] # port on RPI B+ ==> # ls -l  /dev/tty*
  baud_rate = 9600

  # Initialisation of serial connection
  i = 0
  while(True):
    try:
      port = "/dev/tty" + ports[i]
      print(f"Essaie de connexion au port série {ports[i]}...")
      arduino = serial.Serial(port, baud_rate, timeout=1)
      print(f"Connecté à l'Arduino sur le port {ports[i]} à {baud_rate} bauds")
      break
    except serial.SerialException as e:
      print(f"Erreur de connexion : {e}\n")
    i += 1
    if i >= len(ports) :
      i = 0
      time.sleep(5)
  return arduino

def Serial_Synchronisation(arduino: serial.Serial) -> None: # Envoie un message de synchronisation à l'Arduino et attend la réponse.
  while True:
    # Envoi du message
    cmd_to_send = Comms.To_Arduino.RASPBERRY_READY
    print(f"Envoi du message de synchronisation : {cmd_to_send.name}")
    Send_To_Arduino(arduino, cmd_to_send.value)

    time.sleep(1)

    # Réception d’un message
    message = Recv_From_Arduino(arduino)
    if message is None:
      continue  # no complete message yet
    else :
      if message == Comms.From_Arduino.ARDUINO_READY.value:
        cmd_received = Comms.From_Arduino(message)  # conversion int -> enum
        print(f"Réception du message de synchronisation : {cmd_received.name}")
        break

def Send_To_Arduino(arduino: serial.Serial, cmd: int, payload: str = "") -> None:
    if payload:
        message = b'<' + bytes([cmd]) + b'|' + payload.encode("utf-8") + b'>'
    else:
        message = b'<' + bytes([cmd]) + b'>'
    print("Msg sent :", message)  # debug
    arduino.write(message)


def Recv_From_Arduino(arduino:serial) -> int | None: # Receive data from the Arduino (Raw or with start and end marker)

  # wait for the start character
  while True:
    incoming = arduino.read(1)  # lit 1 octet brut
    if not incoming:
      return None
    if incoming == b'<':
      break

  # lire le prochain octet = la commande
  cmd_byte = arduino.read(1)
  #print(f'cmd_byte :{cmd_byte}')
  if not cmd_byte:
    raise TimeoutError("Aucun octet de commande reçu")
  cmd_val = cmd_byte[0]  # convertir byte -> int
  print(f'Cmd receved :{hex(cmd_val)}')

  # attendre le marqueur de fin '>' (ASCII 62)
  while True:
    incoming = arduino.read(1)
    if not incoming:
      continue
    if incoming == b'>':
      break

  return cmd_val

## Excel file usage
# memo
# value of the cell             : print('The value in cell A1 is: '+ws['A1'].value)
# nbr of row and nbr of columns : print('Total number of rows: '+str(ws.max_row)+'. And total number of columns: '+str(ws.max_column))
# writing to a cell             : ws.cell(row=1, column=11, value = 'Sum of Sales') # selection and writing the cell
# save the written  movement       : wb.save(programFileName)  # save the changes --> PERMISSION DENIED --> To see...
def Open_xlsx_WorkBook(programFileName:str) -> openpyxl.Workbook: # Open an excel workbook, mount it, and return it
  print(f"Ouverture du fichier de programmes {programFileName}...")
  workBook = openpyxl.load_workbook(programFileName, data_only=True)  # mounting the workBook
  return workBook
def Program_Choice(workBook:openpyxl.Workbook) -> openpyxl.worksheet.worksheet.Worksheet: #List the programs in the workbook and let the user choose one
  while (True) :
    print("\nQuel programme souhaitez vous utiliser :")
    for idx, x in enumerate(workBook.sheetnames):
      if idx > 1 :
        print(f"{idx-1} : {x}")
    #for idx, x in enumerate(xs):
    #  print(idx, x)

    try :
      choice = int(input("\nVotre choix : "))
      if isinstance(choice, int) :
        if choice > 0 and choice < idx  :
          break
    except ValueError or UnboundLocalError or choice <= 0 or choice >= idx-1:
      print(f"Choix invalide, veuillez faire un choix entre 0 et {idx-1}.")

  workSheets = workBook.worksheets[choice+1] # workSheet
  print(f"Tu as choisis le programme : {workSheets}\n")
  return workSheets
def Read_cell_value(workSheets:openpyxl.worksheet.worksheet.Worksheet, row:int, column:int) -> str | None:  # Read the value of a cell in the worksheet
  # careful --> if empty cells : will return 'None' as a string
  return str(workSheets.cell(row = row, column = column).value)

## movements
# Data treatment
def Speed_percentage_to_us(speedPercentage:float) -> float | None: # Convert speed from percentage to microseconds
  # Convert speed from % to robot compatible speed
  # Delay needs to be 0.0015s = 1500 microseconds at least to protect the motors
  # Risk of damage for the motors if less than 1500 us

  MAX_SPEED = 1500  # max speed in us (lower value = faster)
  MIN_SPEED = 6000  # min speed in us (higher value = slower)

  if speedPercentage < 1.00 or speedPercentage > 100.00 :
    print(f"Vitesse incorrect : {speedPercentage} % - doit être compris entre 1.00 et 100.00 %")
    return None # invalid speed percentage
  else :
    speedUs = np.interp(speedPercentage, [1.0, 100.0], [MIN_SPEED, MAX_SPEED])
    return round(speedUs, 2)
def Gripper_percentage_to_step(gripperPercentage:float) -> float | None: # Convert gripper from percentage to step
  # 0 % = 0 step
  # 100 % = -6000 step
  MAX_GRIPPER_OPENING = -6000 # in step
  MIN_GRIPPER_OPENING = 0 # in step

  if gripperPercentage < 0.0 or gripperPercentage > 100.0:
    print(f"Pourcentage d'ouverture de la pince incorrect : {gripperPercentage} % - doit être compris entre 0.00 et 100.00 %")
    return None # invalid gripper percentage
  else:
    gripperSteps = np.interp(gripperPercentage, [0.0, 100.0], [MIN_GRIPPER_OPENING, MAX_GRIPPER_OPENING])
    return round(gripperSteps, 2) # In steps

def Validate_Move(motorAngles:list[float]) -> int:  # Test if the angles are valid (physical constraints) for the robot's motors  # 0 == valid
  VALID = 0x00
  INVALID = 0xA0
  ERROR_PROGRAM = 0xFF
  
  FORWARD   = 1
  BACKWARDS = 0

  ANGLES_LIMITS = [
    #
    # Motor coders limits in angles
    # considering the robot straight up :
    #
    # Backward  /  Forward
    #    -180   /   +160
    #    -75    /   +135
    #    +90    /   -135
    #    -90    /   +90
    #       illimited
    #    -6000  /   0
    #
    # BW    FW
    [180, 160],   # MOTOR_0 ( in ° )
    [75,  135],   # MOTOR_1
    [90,  135],   # MOTOR_2
    [90,  90 ],   # MOTOR_3
    [0,   0  ],   # MOTOR_4
    [0,   100]    # MOTOR_5 /*closed  opened*/ ( in % 0-100) // 0% = 0 -- 100% = -6000
  ]

  def Angle_calc(angle1:float, angle2:float) -> float:
    theta1 = (180 - 90 - angle1)
    theta2 = (angle2 - 90)
    return (180 - theta2 - theta1)

  for motor_no in Youpi.Motors :
    match motor_no :

      case Youpi.Motors.M0_BASE:
        if( motorAngles[Youpi.Motors.M0_BASE.value] > ANGLES_LIMITS[Youpi.Motors.M0_BASE.value][FORWARD]
          or
          motorAngles[Youpi.Motors.M0_BASE.value] < -ANGLES_LIMITS[Youpi.Motors.M0_BASE.value][BACKWARDS]) :
          return INVALID + Youpi.Motors.M0_BASE.value

      case Youpi.Motors.M1_SHOULDER:
        if( motorAngles[Youpi.Motors.M1_SHOULDER.value] > ANGLES_LIMITS[Youpi.Motors.M1_SHOULDER.value][FORWARD]
          or
          motorAngles[Youpi.Motors.M1_SHOULDER.value] < -ANGLES_LIMITS[Youpi.Motors.M1_SHOULDER.value][BACKWARDS]) :
          return INVALID + Youpi.Motors.M1_SHOULDER.value

      case Youpi.Motors.M2_ELBOW :
        angle = Angle_calc(motorAngles[Youpi.Motors.M1_SHOULDER.value], motorAngles[Youpi.Motors.M2_ELBOW.value])
        
        if( angle < 180 - ANGLES_LIMITS[Youpi.Motors.M2_ELBOW.value][FORWARD]
          or
          360 - angle < ANGLES_LIMITS[Youpi.Motors.M2_ELBOW.value][BACKWARDS]) :
          return INVALID + Youpi.Motors.M2_ELBOW.value

      case Youpi.Motors.M3_PITCH:
        angle = Angle_calc(motorAngles[Youpi.Motors.M2_ELBOW.value], motorAngles[Youpi.Motors.M3_PITCH.value])

        if( angle < 180 - ANGLES_LIMITS[Youpi.Motors.M3_PITCH.value][FORWARD]
          or
          360 - angle < ANGLES_LIMITS[Youpi.Motors.M3_PITCH.value][BACKWARDS]) :
          return INVALID + Youpi.Motors.M3_PITCH.value

      case Youpi.Motors.M4_ROLL:
        # illimited
        pass

      case Youpi.Motors.M5_GRIPPER:
        if( motorAngles[Youpi.Motors.M5_GRIPPER.value] > ANGLES_LIMITS[Youpi.Motors.M5_GRIPPER.value][FORWARD]
          or
          motorAngles[Youpi.Motors.M5_GRIPPER.value] < -ANGLES_LIMITS[Youpi.Motors.M5_GRIPPER.value][BACKWARDS]) :
          return INVALID + Youpi.Motors.M5_GRIPPER.value

      case _ :
        return ERROR_PROGRAM

  return VALID
def Inverse_Kinematic(DH_params:dict, X:float, Y:float, Z:float, pitch:float, roll:float, penOffsetV:float, penOffsetH:float) -> list[float]: # Does the inverse kinematic calculation for the robot's arm

  # the DH_params dictionnary parameters needs to be declared like so :
  # DH_params = {
  #   "D"     : [L1   , 0.0 , 0.0 , 0.0  , L4  , 0.0], #[5] = penOffsetV
  #   "A"     : [0.0  , L2  , L3  , 0.0  , 0.0 , 0.0], #[5] = penOffsetH
  #   "ALPHA" : [90.0 , 0.0 , 0.0 , 90.0 , 0.0 , 0.0]
  # }
  d1 = DH_params["D"][0]
  d2 = DH_params["D"][1] 
  d3 = DH_params["D"][2] 
  d4 = DH_params["D"][3] 
  d5 = DH_params["D"][4]
  d6 = DH_params["D"][5] + penOffsetV

  a1 = DH_params["A"][0]
  a2 = DH_params["A"][1] 
  a3 = DH_params["A"][2]
  a4 = DH_params["A"][3]
  a5 = DH_params["A"][4]
  a6 = DH_params["A"][5] + penOffsetH 

  alpha1 = DH_params["ALPHA"][0]
  alpha2 = DH_params["ALPHA"][1]
  alpha3 = DH_params["ALPHA"][2]
  alpha4 = DH_params["ALPHA"][3]
  alpha5 = DH_params["ALPHA"][4]
  alpha6 = DH_params["ALPHA"][5]

  PD = [X, Y, Z]

  pitch = pitch * math.pi / 180.0  # convert to radians

  RotY = [
    [math.cos(pitch), 0, math.sin(pitch)],
    [0, 1, 0],
    [-math.sin(pitch), 0, math.cos(pitch)] 
  ]

  Pn = [
    RotY[0][0] * 0 + RotY[0][2] * d6,
    0,
    RotY[2][0] * 0 + RotY[2][2] * d6 
  ]

  t = math.atan2(Y, X)

  Pn[0] = PD[0] - Pn[0] * math.cos(t)
  Pn[1] = PD[1] - Pn[1] * math.sin(t)
  Pn[2] = PD[2] - Pn[2]
  
  X = Pn[0]
  Y = Pn[1]
  Z = Pn[2]

  # Inverse Kinematics
  #t1 = (fabs(y) < 1e-5 && fabs(x) < 1e-5) ? 0 : atan2(y, x);
  if (abs(Y) < 1e-5 and abs(X) < 1e-5) :
    t1 = 0
  else :
    t1 = math.atan2(Y, X)

  # This will make the robot to follow actual rotation configuration
  # If removed, then positive means robot end-effector always pointing outside
  # Negative means robot end-effector always pointing inside
  if (abs(X) > 1e-5) :
    #pitch = SIGN(X) * pitch
    pass

  Rn = math.sqrt(X * X + Y * Y) - L4 * math.sin(pitch)
  Zn = Z - L4 * math.cos(pitch) - L1

  C3 = (Rn * Rn + Zn * Zn - L2 * L2 - L3 * L3) / (2 * L2 * L3)
  C3 = min(1, max(C3, -1))

  t3 = -math.acos(C3)
  if (pitch < 0) :
    t3 = -t3

  t2 = math.atan2(Zn, Rn) - math.atan2(L3 * math.sin(t3), L2 + L3 * math.cos(t3))
  t4 = -pitch - t2 - t3 + math.pi
  t5 = roll

  # Motor Compatable Angles */
  t1_Motor = t1
  t2_Motor = t2 - math.pi/2
  t3_Motor = t2_Motor + t3
  t4_Motor = t3_Motor + t4 - math.pi/2
  t5_Motor = t5 - t4_Motor
    
  #t1_Motor = -t1_Motor 
  t2_Motor = -t2_Motor
  t3_Motor = -t3_Motor
  t4_Motor = -t4_Motor
  t5_Motor = -t5_Motor

  # Final Angles */
  motorAngles = [0.0] * 5

  motorAngles[0] = t1_Motor * 180 / math.pi
  motorAngles[1] = t2_Motor * 180 / math.pi
  motorAngles[2] = t3_Motor * 180 / math.pi
  motorAngles[3] = t4_Motor * 180 / math.pi
  motorAngles[4] = t5_Motor * 180 / math.pi

  motorAngles = [round(x, 3) for x in motorAngles] # Round the angles to 3 decimal places

  return motorAngles

def Get_movements(workSheets:openpyxl.worksheet.worksheet.Worksheet, row_no:int) -> list | None: # Get the data of the movement from the worksheet at the given row number - return everything as float except the first element (movement mode) as string

  if row_no < Excel_Params.STARTING_ROW :
    print(f"Erreur de lecture de la ligne {row_no} : la ligne doit être supérieure ou égale à {Excel_Params.STARTING_ROW}")
    return None

  # Get the movement mode of the movement, wich determine how much data to read
  movementMode  =  Read_cell_value(workSheets, row_no, column=1)
  match movementMode :
    case Youpi.movement_Mode.JOINT.value :
      startColumn = 11
      nbrOfData   = 5
    case Youpi.movement_Mode.LINEAR.value :
      startColumn = 4
      nbrOfData   = 7
    case _ :  # invalid
      print(f"Choix de mouvement erroné - valeur de la ligne {row_no} : mouvementMode = {movementMode}")
      return None # error

  # Get the speed, check if valid and convert to Us
  speed = Read_cell_value(workSheets, row_no, column=2)
  try :
    speed = float(speed)
  except ValueError :
    print(f"Erreur de lecture de la vitesse du mouvement à la ligne {row_no} : speed = {speed} - doit être un nombre")
    return None
  
  # Get the gripper percentage, check if valid and convert to step
  gripperPercentage = Read_cell_value(workSheets, row_no, column=3)
  try :
    gripperPercentage = float(gripperPercentage)
  except ValueError :
    print(f"Erreur de lecture du pourcentage de fermeture de la pince à la ligne {row_no} : gripperPercentage = {gripperPercentage} - doit être un nombre")
    return None
  
  movements = [] # List to store the data of the movement
  movements.append(movementMode)
  movements.append(speed)
  movements.append(gripperPercentage)
 
  # Read the movement data from the worksheet and add them to the list
  for i in range(startColumn, startColumn + nbrOfData):
    movement = Read_cell_value(workSheets, row_no, column=i)
    try :
      movement = float(movement)
    except ValueError :
      print(f"Erreur de lecture de la valeur de la cellule à la ligne {row_no} et colonne {i} : cell_value = {movement} - doit être un nombre")
      return None
    movements.append(movement)
  return  movements
def Get_Program(workSheets:openpyxl.worksheet.worksheet.Worksheet) -> list | None:  # Read all the row of a program to create a 2d array of it

  # Gets the number of movements the program contains
  nbrOfmovements = Read_cell_value(workSheets, row=2, column=1)
  try :
    nbrOfmovements = int(nbrOfmovements)
  except ValueError :
    print(f"Erreur de lecture du nombre de mouvements à exécuter : {nbrOfmovements} - doit être un nombre entier")
    return None
  if nbrOfmovements <= 0 :
    print("Aucun mouvement à exécuter.")
    return None

  program = [] # 2d list containing all the movements of the program
  rowNbr = 0
  
  while (True) :
    if len(program) >= nbrOfmovements : # if we filled up the array with all the movements
      break

    # Get the movement data from a dedicated row of the prograù
    movements = Get_movements(workSheets, Excel_Params.STARTING_ROW + rowNbr)
    if movements is None :
      return None # error

    program.append(movements)
    #print(f"movement data in row no {Excel_Params.STARTING_ROW + rowNbr} : {movementsDatas}")
    rowNbr += 1
  
  #print('All movements of the selected worksheet : ')
  #for x in movementsDatas : print(f'{x}')
  return program

def Convert_movement_to_joint(movements:list) -> list[float]: # Convert the program data to motor angles
   
  movementType  = movements[0]

  speed         = movements[1]
  speedUs = Speed_percentage_to_us(speed)
  if speedUs is None :
    return None # error
  
  gripperPercentage  = movements[2]
  gripperSteps = Gripper_percentage_to_step(gripperPercentage)
  if gripperSteps is None :
    return None # error

  if movementType == Youpi.movement_Mode.LINEAR.value :
    motorAnglesTargets = Inverse_Kinematic( # return the angles of the motors (j0 - j4) in degrees
      DH_params   = Youpi.DH_Params,
      X           = movements[3],
      Y           = movements[4],
      Z           = movements[5],
      pitch       = movements[6],
      roll        = movements[7],
      penOffsetV  = movements[8],
      penOffsetH  = movements[9]
    )
    #print(f"Inverse kinematic angles : {motorAnglesTargets}")
  elif movementType == Youpi.movement_Mode.JOINT.value :
    j0 = movements[3]
    j1 = movements[4]
    j2 = movements[5]
    j3 = movements[6]
    j4 = movements[7]

    motorAnglesTargets = [j0, j1, j2, j3, j4]

  mvtJointDatas = [speedUs] + motorAnglesTargets + [gripperSteps]
  
  return mvtJointDatas


    # -------------------------------------------------------------------------------------------------------------
    # Continuer ici --> convertir les angles en pas moteurs ici ou créer une fonction de conversion


    #print(f"mouvement récupéré ligne {Excel_Params.STARTING_ROW + rowNbr - 1} : {motorAnglesTargets}")

    #errorID = Validate_Move(motorAnglesTargets)
    #if errorID != 0x00 :
    #  print(f"Erreur de validation des angles du mouvement à la ligne {Excel_Params.STARTING_ROW + rowNbr - 1} : errorID = {hex(errorID)}")
    #  print(f"Angles invalides : {errorID - 0xA0} = {motorAnglesTargets[errorID - 0xA0]}")
    #  return errorID
    #else :
      #motorAnglesTargets = [str(x) for x in motorAnglesTargets]

      # Add the movement datas to the list to send to the arduino

# Important functions
def Program_Execution(arduino:serial, workSheets:openpyxl.worksheet.worksheet.Worksheet) -> int:  # Execute the program by sending the movements datas to the arduino

  # Gets all the movements datas from the worksheet
  program = Get_Program(workSheets)
  if program == None :
    return None

  print(f"program execution value :{Comms.To_Arduino.PROGRAM_EXECUTION.value}")
  # sends the message to the arduino to be ready to receive the movements datas
  Send_To_Arduino(arduino, Comms.To_Arduino.PROGRAM_EXECUTION.value)

  # Execution of the movements
  mvtNbr = 0
  while mvtNbr < len(program) :
    # wait for the arduino to be ready to receive the movements datas
    while (True) :
      cmd = Recv_From_Arduino(arduino)
      if cmd == Comms.From_Arduino.READY_TO_RECEIVE.value :
        break
      #elif :
        # TIMEOUT ?

    # convert from raw XLS format to joint format (Speed, J1, J2, ..., J5, gripper)
    print(f"Mouvement no {mvtNbr+1} sur {len(program)} : {program[mvtNbr]}")
    movementJoint = Convert_movement_to_joint(program[mvtNbr])
    print(f"mouvement no {mvtNbr+1} sur {len(program)} convertis : {movementJoint}")
    
    # create the compatible movement string to send to the arduino
    separator = "_"
    #movementString = separator.join(str(int(round(x))) for x in movementJoint)
    movementString = separator.join(str(x) for x in movementJoint)
    print(f"string joint : {movementString}")

    #if abortProgram == True :
    #Send_To_Arduino(arduino, Comms.To_Arduino.ABORT_PROGRAM.value, direct=False)

    # send the movement datas to the arduino
    Send_To_Arduino(arduino, Comms.To_Arduino.START_MOVEMENT.value, movementString)
    
    # wait for the checksum of the movement - checking if the good data was received
    # Calculate the checksum of the movement
    #while (True) :
    #  receivedData = Recv_From_Arduino(arduino, direct=False) # receive the message from the arduino
    #  if receivedData == MvtChecksum :
    #    break

    # wait for the arduino to finish the movement
    while (True) :
      #Send_To_Arduino(arduino, '1') # send a constant value to the arduino to indicate that no abort is requested
   
      cmd = Recv_From_Arduino(arduino) # receive the message from the arduino
      if cmd == Comms.From_Arduino.MOVEMENT_FINISHED.value :
        break
    mvtNbr = mvtNbr + 1

  Send_To_Arduino(arduino, Comms.To_Arduino.PROGRAM_FINISHED.value)

  return 0x00 # success


# --------------------- Main ---------------------

#from pynput.keyboard import Key, Controller #type:ignore

def main() -> None:
  arduino = Serial_Initialisation()
  print("Connecté à l'Arduino, prêt à recevoir des commandes.")
  print("Appuyez sur Ctrl+C pour quitter le programme.")

  try:
    while True:

      # Menu pour envoyer différentes données
      print("\nChoisissez une option pour envoyer des données :")
      print("0 - Quitter")
      print("1 - choix programme + execution")

      choice = input("Votre choix : ")

      match choice :
        case '0' :
          print("Déconnexion...")
          break   # Break the while loop, ending the program
        case '1' :
          workBook = Open_xlsx_WorkBook(Excel_Params.PROGRAM_FILE_NAME)
          workSheets = Program_Choice(workBook)
          Program_Execution(arduino, workSheets)
          continue
        case _ :
          print("Choix invalide.")
          continue
        
  except KeyboardInterrupt:
    print("\nDéconnexion en cours...")
  finally:
    arduino.close()  # Fermer la connexion série
    print("Déconnecté.")


if __name__ == "__main__":
  main()

#while True:
#  #keyboard = Controller()
#  key = input()
#
#  if (key == 32) :
#    while (key == 32):
#      arduino.write((1)).encode()
#      key = input()
#    arduino.write((0)).encode()
#  else :
#    arduino.write((0)).encode()
#
#  #if keyboard.press(Key.space) :
#  #  while ( not keyboard.release(Key.space)) :
#  #    arduino.write((1 + "\n").encode())
#  #  arduino.write((0 + "\n").encode())