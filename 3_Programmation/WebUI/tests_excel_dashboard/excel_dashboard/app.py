
from flask import Flask, render_template, request, redirect, url_for
import openpyxl
import os

app = Flask(__name__)
EXCEL_FILE = "data.xlsx"

@app.route("/", methods=["GET", "POST"])
def index():
    wb = openpyxl.load_workbook(EXCEL_FILE)
    sheet_names = wb.sheetnames
    selected_sheet = request.args.get("sheet", sheet_names[0])
    sheet = wb[selected_sheet]
    data = [[cell.value for cell in row] for row in sheet.iter_rows()]

    if request.method == "POST":
        row = int(request.form["row"])
        col = int(request.form["col"])
        value = request.form["value"]
        sheet.cell(row=row+1, column=col+1, value=value)
        wb.save(EXCEL_FILE)
        return redirect(url_for("index", sheet=selected_sheet))

    return render_template("index.html", data=data, sheet_names=sheet_names, selected_sheet=selected_sheet)

if __name__ == "__main__":
    app.run(host="0.0.0.0", port=5000)
