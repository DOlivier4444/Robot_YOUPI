from flask import Flask, render_template, request, redirect, url_for
from openpyxl import load_workbook
import os

app = Flask(__name__)

EXCEL_FILE = os.path.join(os.path.dirname(__file__), "data.xlsx")

@app.route("/", methods=["GET", "POST"])
def index():
    wb = load_workbook(EXCEL_FILE)
    sheet_names = wb.sheetnames
    selected_sheet = request.form.get("sheet") or request.args.get("sheet") or sheet_names[0]
    sheet = wb[selected_sheet]

    if request.method == "POST" and "update_cell" in request.form:
        row = int(request.form["row"])
        col = int(request.form["col"])
        value = request.form["value"]
        sheet.cell(row=row, column=col).value = value
        wb.save(EXCEL_FILE)
        return redirect(url_for("index", sheet=selected_sheet))

    data = [[cell.value for cell in row] for row in sheet.iter_rows()]
    return render_template("index.html", data=data, sheets=sheet_names, active_sheet=selected_sheet)

if __name__ == "__main__":
    app.run(host="0.0.0.0", port=5000, debug=True)
