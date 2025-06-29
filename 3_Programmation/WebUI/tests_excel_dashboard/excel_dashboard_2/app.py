from flask import Flask, render_template, request, jsonify
import openpyxl
import os

app = Flask(__name__)

EXCEL_PATH = "data.xlsx"

@app.route("/")
def index():
    return render_template("index.html")

@app.route("/sheets")
def get_sheets():
    wb = openpyxl.load_workbook(EXCEL_PATH)
    return jsonify(wb.sheetnames)

@app.route("/sheet/<name>")
def get_sheet_data(name):
    wb = openpyxl.load_workbook(EXCEL_PATH)
    ws = wb[name]
    data = [[cell.value for cell in row] for row in ws.iter_rows()]
    return jsonify(data)

@app.route("/update", methods=["POST"])
def update_cell():
    content = request.json
    wb = openpyxl.load_workbook(EXCEL_PATH)
    ws = wb[content["sheet"]]
    row = int(content["row"])
    col = int(content["col"])
    ws.cell(row=row, column=col).value = content["value"]
    wb.save(EXCEL_PATH)
    return jsonify(success=True)

if __name__ == "__main__":
    if not os.path.exists(EXCEL_PATH):
        wb = openpyxl.Workbook()
        wb.save(EXCEL_PATH)
    app.run(debug=True)
