from enum import Enum
import os
import serial #type:ignore
import openpyxl #https://www.datacamp.com/tutorial/python-excel-tutorial  #type:ignore
import time
import math


# --------------------- Classes ---------------------
## 
class Excel_Params :
  PROGRAM_FILE_NAME = 'programs.xlsx'
  STARTING_ROW = 4 # row where the program start in the xlsx file

## Communications
class Comms :
  START_MARKER = '<'
  END_MARKER   = '>'


  class To_Arduino(str, Enum) :
    READY_TO_RECEIVE  = 'Ready to receive'
    RASPBERRY_READY   = 'Raspberry ready!'
    
    START_MOVEMENT    = 'Start movement'    #used later for handshake of mvt data ??

    PROGRAM_EXECUTION = 'Program execution'
    PROGRAM_FINISHED  = 'Program finished'
    ABORT_PROGRAM     = 'Abort program'


  class From_Arduino(str, Enum) :
    READY_TO_RECEIVE  = 'Ready to receive'
    ARDUINO_READY     = 'Arduino ready!'

    MOVEMENT_FINISHED = 'Movement finished'


  class General(str, Enum) :
    ERROR_MOVEMENT    = 'Error movement !'

## Robot's parameters
L1 = 0.28
L2 = 0.162
L3 = 0.162
L4 = 0.15
class Youpi :
  class Motors(Enum) :
    M0_BASE     = 0  # motor0 - base
    M1_SHOULDER = 1  # motor1 - shoulder
    M2_ELBOW    = 2  # motor2 - elbow
    M3_PITCH    = 3  # motor3 - wrist pitch
    M4_ROLL     = 4  # motor4 - wrist roll
    M5_GRIPPER  = 5  # motor5 - gripper

  class Robot_Arms(Enum) :
    # Lengths of the Robot's arm
    L1 = L1
    L2 = L2
    L3 = L3
    L4 = L4

  DH_Params = {
    # DH parameters
    #
    # const float D1 = L1             ; const float A1 = 0.0            ; const float ALPHA1 = 90.0 ;
    # const float D2 = 0.0            ; const float A2 = L2             ; const float ALPHA2 = 0.0  ;
    # const float D3 = 0.0            ; const float A3 = L3             ; const float ALPHA3 = 0.0  ;
    # const float D4 = 0.0            ; const float A4 = 0.0            ; const float ALPHA4 = 90.0 ;
    # const float D5 = L4             ; const float A5 = 0.0            ; const float ALPHA5 = 0.0  ;
    # const float D6 = penOffsetV     ; const float A6 = penOffsetH     ; const float ALPHA6 = 0.0  ;
    "D"     : [L1   , 0.0 , 0.0 , 0.0  , L4  , 0.0], #[5] = penOffsetV
    "A"     : [0.0  , L2  , L3  , 0.0  , 0.0 , 0.0], #[5] = penOffsetH
    "ALPHA" : [90.0 , 0.0 , 0.0 , 90.0 , 0.0 , 0.0]
  }

  class Movement_Mode(str, Enum) :
    JOINT = "J"
    LINEAR = "L"

# --------------------- Functions ---------------------
def Serial_Initialisation() -> serial :
  arduino = Serial_Connexion()
  Serial_Synchronisation(arduino)
  return arduino

# Serial communication
def Serial_Connexion() -> serial:
  # Configuration of serial connection
  ports = ["ACM0", "ACM1", "USB0", "USB1"] # port on RPI B+ ==> # ls -l  /dev/tty*
  baud_rate = 9600

  # Initialisation of serial connection
  i = 0
  while(True):
    try:
      port = "/dev/tty" + ports[i]
      print(f"Essaie de connexion au port série {ports[i]}...")
      arduino = serial.Serial(port, baud_rate, timeout=1)
      print(f"Connecté à l'Arduino sur le port {ports[i]} à {baud_rate} bauds")
      break
    except serial.SerialException as e:
      print(f"Erreur de connexion : {e}\n")
    i += 1
    if i >= len(ports) :
      i = 0
      time.sleep(5)
  return arduino
def Serial_Synchronisation(arduino:serial) -> None:
  while(True) :
    message = Comms.To_Arduino.RASPBERRY_READY.value
    print(f"Envoie du message de synchronisation : " + message)
    Send_To_Arduino(arduino, message, direct=False)
    
    message = Recv_From_Arduino(arduino, direct=False)
    if message == Comms.From_Arduino.ARDUINO_READY.value :
      print(f"Réception du message de synchronisation : " + message)
      break
def Send_To_Arduino(arduino:serial, data:str, direct:bool = False) -> None:
  if not direct :
    # add start and end marker to the data
    data = Comms.START_MARKER + data + Comms.END_MARKER
  arduino.write(bytes(data, 'utf-8'))
def Recv_From_Arduino(arduino:serial, direct:bool = False) -> str:
  if direct :
    # read the data directly without start and end marker
    return arduino.read().decode('utf-8')
  
  incomingChar = ""
  receivedData = ""

  # wait for the start character
  while incomingChar != Comms.START_MARKER :
    incomingChar = arduino.read().decode('utf-8')

  # save data until the end marker is found
  while(True) :
    incomingChar = arduino.read().decode('utf-8')
    if incomingChar == Comms.END_MARKER :
      break
    else :
      receivedData = receivedData + incomingChar
  return receivedData # Without start and endmarker

# Excel file usage
def Open_xlsx_WorkBook(programFileName:str) -> openpyxl.Workbook:
  print(f"Ouverture du fichier de programmes {programFileName}...")
  workBook = openpyxl.load_workbook(programFileName, data_only=True)  # mounting the workBook
  return workBook
def Program_Choice(workBook:openpyxl.Workbook) -> openpyxl.worksheet.worksheet.Worksheet:

  while (True) :
    print("\nQuel programme souhaitez vous utiliser :")
    for idx, x in enumerate(workBook.sheetnames):
      if idx > 1 :
        print(f"{idx-1} : {x}")
    #for idx, x in enumerate(xs):
    #  print(idx, x)

    choice = int(input("\nVotre choix : "))
    if choice > 0 and choice <= idx :
      break
    else :
      print("Choix invalide, veuillez choisir un programme présent dans la liste.")

  workSheets = workBook.worksheets[choice+1] # workSheet
  print(f"Tu as choisis le programme : {workSheets}\n")
  return workSheets
def Read_cell_value(workSheets:openpyxl.worksheet.worksheet.Worksheet, row:int, column:int) -> str | None: # careful ! --> if read none, it will return 'None' as a string... to see later
  return str(workSheets.cell(row = row, column = column).value)
# memo
# value of the cell             : print('The value in cell A1 is: '+ws['A1'].value)
# nbr of row and nbr of columns : print('Total number of rows: '+str(ws.max_row)+'. And total number of columns: '+str(ws.max_column))
# writing to a cell             : ws.cell(row=1, column=11, value = 'Sum of Sales') # selection and writing the cell
# save the written values       : wb.save(programFileName)  # save the changes --> PERMISSION DENIED --> To see...

## Movements
# Data treatment
def Speed_percentage_to_us(speedPercentage:float) -> float:  
  # Convert speed from % to robot speed
  MAX_SPEED = 1500 # in us                      #Delay needs to be 0.0015s = 1500 microseconds at least to protect the motors // Danger for the motors if less than 1500 us, using this speed is not recommanded
  MIN_SPEED = 6000 # in us

  if speedPercentage >= 1.0 and speedPercentage <= 100.0 :
    a = (MAX_SPEED - MIN_SPEED) / (100.0 - 1.0)
    b = MAX_SPEED - a * 100.0
    speedUs = a * speedPercentage + b
    return int(speedUs) # return speed in us
  else :
    speedUs = 0
    return speedUs
def Gripper_percentage_to_step(gripperPercentage:float) -> float:
  # 0% = 0 -- 100% = -6000
  MAX_GRIPPER_OPENING = -6000 # in step
  MIN_GRIPPER_OPENING = 0 # in step

  if gripperPercentage >= 0.0 and gripperPercentage <= 100.0 :
    a = (MAX_GRIPPER_OPENING - MIN_GRIPPER_OPENING) / (100.0 - 1.0)
    b = MAX_GRIPPER_OPENING - a * 100.0
    gripperStep = a * gripperPercentage + b
    return int(gripperStep) # return speed in us
  else :
    gripperStep = 0
    return gripperStep

def Validate_Move(motorAngles:list[float]) -> int: # 0 == valid
  FORWARD   = 1
  BACKWARDS = 0

  ANGLES_LIMITS = [
    #
    # Motor coders limits in angles
    # considering the robot straight up :
    #
    # Backward  /  Forward
    #    -180   /   +160
    #    -75    /   +135
    #    +90    /   -135
    #    -90    /   +90
    #       illimited
    #    -6000  /   0
    #
    # BW    FW
    [180, 160],   # MOTOR_0 ( in ° )
    [75,  135],   # MOTOR_1
    [90,  135],   # MOTOR_2
    [90,  90 ],   # MOTOR_3
    [0,   0  ],   # MOTOR_4
    [0,   100]    # MOTOR_5 /*closed  opened*/ ( in % 0-100) // 0% = 0 -- 100% = -6000
  ]

  def Angle_calc(angle1:float, angle2:float) -> float:
    theta1 = (180 - 90 - angle1)
    theta2 = (angle2 - 90)
    return (180 - theta2 - theta1)

  for motor_no in Youpi.Motors :
    match motor_no :

      case Youpi.Motors.M0_BASE:
        if( motorAngles[Youpi.Motors.M0_BASE.value] > ANGLES_LIMITS[Youpi.Motors.M0_BASE.value][FORWARD]
          or
          motorAngles[Youpi.Motors.M0_BASE.value] < -ANGLES_LIMITS[Youpi.Motors.M0_BASE.value][BACKWARDS]) :
          return 0xA0 + Youpi.Motors.M0_BASE.value

      case Youpi.Motors.M1_SHOULDER:
        if( motorAngles[Youpi.Motors.M1_SHOULDER.value] > ANGLES_LIMITS[Youpi.Motors.M1_SHOULDER.value][FORWARD]
          or
          motorAngles[Youpi.Motors.M1_SHOULDER.value] < -ANGLES_LIMITS[Youpi.Motors.M1_SHOULDER.value][BACKWARDS]) :
          return 0xA0 + Youpi.Motors.M1_SHOULDER.value

      case Youpi.Motors.M2_ELBOW :
        angle = Angle_calc(motorAngles[Youpi.Motors.M1_SHOULDER.value], motorAngles[Youpi.Motors.M2_ELBOW.value])
        
        if( angle < 180 - ANGLES_LIMITS[Youpi.Motors.M2_ELBOW.value][FORWARD]
          or
          360 - angle < ANGLES_LIMITS[Youpi.Motors.M2_ELBOW.value][BACKWARDS]) :
          return 0xA0 + Youpi.Motors.M2_ELBOW.value

      case Youpi.Motors.M3_PITCH:
        angle = Angle_calc(motorAngles[Youpi.Motors.M2_ELBOW.value], motorAngles[Youpi.Motors.M3_PITCH.value])

        if( angle < 180 - ANGLES_LIMITS[Youpi.Motors.M3_PITCH.value][FORWARD]
          or
          360 - angle < ANGLES_LIMITS[Youpi.Motors.M3_PITCH.value][BACKWARDS]) :
          return 0xA0 + Youpi.Motors.M3_PITCH.value

      case Youpi.Motors.M4_ROLL:
        # illimited
        pass

      case Youpi.Motors.M5_GRIPPER:
        if( motorAngles[Youpi.Motors.M5_GRIPPER.value] > ANGLES_LIMITS[Youpi.Motors.M5_GRIPPER.value][FORWARD]
          or
          motorAngles[Youpi.Motors.M5_GRIPPER.value] < -ANGLES_LIMITS[Youpi.Motors.M5_GRIPPER.value][BACKWARDS]) :
          return 0xA0 + Youpi.Motors.M5_GRIPPER.value

      case _ :
        return 0xFF

  return 0x00
def Inverse_Kinematic(DH_params:dict, X:float, Y:float, Z:float, pitch:float, roll:float, penOffsetV:float, penOffsetH:float) -> list[float]: 

  # the DH_params dictionnary parameters needs to be declared like so :
  # DH_params = {
  #   "D"     : [L1   , 0.0 , 0.0 , 0.0  , L4  , 0.0], #[5] = penOffsetV
  #   "A"     : [0.0  , L2  , L3  , 0.0  , 0.0 , 0.0], #[5] = penOffsetH
  #   "ALPHA" : [90.0 , 0.0 , 0.0 , 90.0 , 0.0 , 0.0]
  # }
  d1 = DH_params["D"][0]
  d2 = DH_params["D"][1] 
  d3 = DH_params["D"][2] 
  d4 = DH_params["D"][3] 
  d5 = DH_params["D"][4]
  d6 = DH_params["D"][5] + penOffsetV

  a1 = DH_params["A"][0]
  a2 = DH_params["A"][1] 
  a3 = DH_params["A"][2]
  a4 = DH_params["A"][3]
  a5 = DH_params["A"][4]
  a6 = DH_params["A"][5] + penOffsetH 

  alpha1 = DH_params["ALPHA"][0]
  alpha2 = DH_params["ALPHA"][1]
  alpha3 = DH_params["ALPHA"][2]
  alpha4 = DH_params["ALPHA"][3]
  alpha5 = DH_params["ALPHA"][4]
  alpha6 = DH_params["ALPHA"][5]

  PD = [X, Y, Z]

  pitch = pitch * math.pi / 180.0  # convert to radians

  RotY = [
    [math.cos(pitch), 0, math.sin(pitch)],
    [0, 1, 0],
    [-math.sin(pitch), 0, math.cos(pitch)] 
  ]

  Pn = [
    RotY[0][0] * 0 + RotY[0][2] * d6,
    0,
    RotY[2][0] * 0 + RotY[2][2] * d6 
  ]

  t = math.atan2(Y, X)

  Pn[0] = PD[0] - Pn[0] * math.cos(t)
  Pn[1] = PD[1] - Pn[1] * math.sin(t)
  Pn[2] = PD[2] - Pn[2]
  
  X = Pn[0]
  Y = Pn[1]
  Z = Pn[2]

  # Inverse Kinematics
  #t1 = (fabs(y) < 1e-5 && fabs(x) < 1e-5) ? 0 : atan2(y, x);
  if (abs(Y) < 1e-5 and abs(X) < 1e-5) :
    t1 = 0
  else :
    t1 = math.atan2(Y, X)

  # This will make the robot to follow actual rotation configuration
  # If removed, then positive means robot end-effector always pointing outside
  # Negative means robot end-effector always pointing inside
  if (abs(X) > 1e-5) :
    #pitch = SIGN(X) * pitch
    pass

  Rn = math.sqrt(X * X + Y * Y) - L4 * math.sin(pitch)
  Zn = Z - L4 * math.cos(pitch) - L1

  C3 = (Rn * Rn + Zn * Zn - L2 * L2 - L3 * L3) / (2 * L2 * L3)
  C3 = min(1, max(C3, -1))

  t3 = -math.acos(C3)
  if (pitch < 0) :
    t3 = -t3

  t2 = math.atan2(Zn, Rn) - math.atan2(L3 * math.sin(t3), L2 + L3 * math.cos(t3))
  t4 = -pitch - t2 - t3 + math.pi
  t5 = roll

  # Motor Compatable Angles */
  t1_Motor = t1
  t2_Motor = t2 - math.pi/2
  t3_Motor = t2_Motor + t3
  t4_Motor = t3_Motor + t4 - math.pi/2
  t5_Motor = t5 - t4_Motor
    
  #t1_Motor = -t1_Motor 
  t2_Motor = -t2_Motor
  t3_Motor = -t3_Motor
  t4_Motor = -t4_Motor
  t5_Motor = -t5_Motor

  # Final Angles */
  motorAngles = [0.0] * 5

  motorAngles[0] = round(t1_Motor * 180 / math.pi, 3)
  motorAngles[1] = round(t2_Motor * 180 / math.pi, 3)
  motorAngles[2] = round(t3_Motor * 180 / math.pi, 3)
  motorAngles[3] = round(t4_Motor * 180 / math.pi, 3)
  motorAngles[4] = round(t5_Motor * 180 / math.pi, 3)

  return motorAngles

def Get_Movement_Datas(workSheets:openpyxl.worksheet.worksheet.Worksheet, row_no:int) -> list[str] | None:

  movementMode  =  Read_cell_value(workSheets, row_no, column=1)

  match movementMode :
    case Youpi.Movement_Mode.JOINT.value :
      startColumn = 11
      nbrOfData   = 5
    case Youpi.Movement_Mode.LINEAR.value :
      startColumn = 4
      nbrOfData   = 7
    case _ :
      print(f"Choix de mouvement erroné - valeur de la ligne {row_no} : mouvementMode = {movementMode}")
      return None # error

  speed = Read_cell_value(workSheets, row_no, column=2)
  speedUs = Speed_percentage_to_us(speed)
  if speedUs is None :
    print(f"Erreur de lecture de la vitesse du mouvement à la ligne {row_no} : speed = {speedUs}")
    return None
  
  gripperPercentage = Read_cell_value(workSheets, row_no, column=3)
  gripperSteps = Gripper_percentage_to_step(gripperPercentage)
  if gripperSteps is None :
    print(f"Erreur de lecture du pourcentage de fermeture de la pince à la ligne {row_no} : gripperPercentage = {gripperSteps}")
    return None
  
  values = []
  values.append(movementMode)
  values.append(speedUs)
  values.append(gripperSteps)
 
  #print(f"start column : {startColumn} \nend column : {startColumn + nbrOfData}")

  for i in range(startColumn, startColumn + nbrOfData):
    cell_value = Read_cell_value(workSheets, row_no, column=i)

    if cell_value is None :
      print(f"Erreur de lecture de la valeur de la cellule à la ligne {row_no} et colonne {i} : cell_value = {cell_value}")
      return None # error
    else :
      values.append( cell_value )
  #print(values)
  return values
def Get_Program_Datas(workSheets:openpyxl.worksheet.worksheet.Worksheet) -> list[str] | int:

  nbrOfMovements = Read_cell_value(workSheets, row=2, column=1)
  if nbrOfMovements is None :
    print("Erreur de lecture du nombre de mouvements à exécuter.")
    return 0xD0
  
  nbrOfMovements = int(float(nbrOfMovements))
  if nbrOfMovements <= 0 :
    print("Aucun mouvement à exécuter.")
    return 0xD2
  #sendToArduino(arduino, str(nbrOfMovements))

  mvtsJointDatas = [] # list of movements datas to send to the arduino
  rowNbr = 0
  
  while (True) :
    if len(mvtsJointDatas) >= nbrOfMovements :
      break

    mvtDatas = Get_Movement_Datas(workSheets, Excel_Params.STARTING_ROW + rowNbr)
    if mvtDatas is None :
      return 0xD4
    rowNbr += 1
    #print(f"Movement data : {mvtDatas}")

    if mvtDatas[0] == Youpi.Movement_Mode.LINEAR.value :
      motorAnglesTargets = Inverse_Kinematic(
        DH_params   = Youpi.DH_Params,
        X           = float(mvtDatas[3]),
        Y           = float(mvtDatas[4]),
        Z           = float(mvtDatas[5]),
        pitch       = float(mvtDatas[6]),
        roll        = float(mvtDatas[7]),
        penOffsetV  = float(mvtDatas[8]),
        penOffsetH  = float(mvtDatas[9])
      )
      motorAnglesTargets.append(float(mvtDatas[2])) # add the gripper % to the list

      #print(f"Inverse kinematic angles : {motorAnglesTargets}")
    else : # joint

      motorAnglesTargets = mvtDatas.copy()
      motorAnglesTargets.pop(0) # Movement mode
      motorAnglesTargets.pop(0) # Speed

      motorAnglesTargets.pop(0) # Gripper percentage
      motorAnglesTargets.append(mvtDatas[2]) # move the gripper percentage to the end of the list
      
      motorAnglesTargets = [float(x) for x in motorAnglesTargets]

    #print(f"mouvement récupéré ligne {Excel_Params.STARTING_ROW + rowNbr - 1} : {motorAnglesTargets}")

    errorID = Validate_Move(motorAnglesTargets)
    if errorID != 0x00 :
      print(f"Erreur de validation des angles du mouvement à la ligne {Excel_Params.STARTING_ROW + rowNbr - 1} : errorID = {hex(errorID)}")
      print(f"Angles invalides : {errorID - 0xA0} = {motorAnglesTargets[errorID - 0xA0]}")
      return errorID
    else :
      motorAnglesTargets = [str(x) for x in motorAnglesTargets]

      speed = mvtDatas[1]
      mvtJointDatas = [speed] + motorAnglesTargets
      mvtJointDatas = '_'.join(mvtJointDatas)

      mvtsJointDatas.append(mvtJointDatas)
      
  #print(f"\nmvt total : ")
  #print(*mvtsJointDatas, sep='\n')
  return mvtsJointDatas
#def Validate_Program_Datas(mvtsJointDatas:list[float]) -> int:

# Important functions
def Program_Execution(arduino:serial, workSheets:openpyxl.worksheet.worksheet.Worksheet) -> int:

  mvtsJointDatas = Get_Program_Datas(workSheets)
  if isinstance(mvtsJointDatas, int) :
    print(f"Erreur lors de la récupération des données du programme : {hex(mvtsJointDatas)}")
    return mvtsJointDatas


  Send_To_Arduino(arduino, Comms.To_Arduino.PROGRAM_EXECUTION.value, direct=False)


  i = 0
  while ( i < len(mvtsJointDatas) ) :

    while (True) :
      receivedData = Recv_From_Arduino(arduino, direct=False)
      if receivedData == Comms.From_Arduino.READY_TO_RECEIVE.value :
        break

    print(f"Envoi du mouvement {i+1} sur {len(mvtsJointDatas)} : {mvtsJointDatas[i]}")

    #if abortProgram == True :
    #Send_To_Arduino(arduino, Comms.To_Arduino.ABORT_PROGRAM.value, direct=False)
    

    #print(mvtsDatas[i])
    Send_To_Arduino(arduino, mvtsJointDatas[i], direct=False) # send the movement datas to the arduino

    while (True) :
      receivedData = Recv_From_Arduino(arduino, direct=False) # receive the message from the arduino
      if receivedData == Comms.From_Arduino.MOVEMENT_FINISHED.value :
        break

    i = i + 1

  Send_To_Arduino(arduino, Comms.To_Arduino.PROGRAM_FINISHED.value, direct=False)

  return 0x00 # success


# --------------------- Main ---------------------

#from pynput.keyboard import Key, Controller #type:ignore

def main() -> None:
  arduino = Serial_Initialisation()
  print("Connecté à l'Arduino, prêt à recevoir des commandes.")
  print("Appuyez sur Ctrl+C pour quitter le programme.")

  try:
    while True:

      # Menu pour envoyer différentes données
      print("\nChoisissez une option pour envoyer des données :")
      print("0 - Quitter")
      print("1 - choix programme + execution")

      choice = input("Votre choix : ")

      match choice :
        case '0' :
          print("Déconnexion...")
          break   # Break the while loop, ending the program
        case '1' :
          workBook = Open_xlsx_WorkBook(Excel_Params.PROGRAM_FILE_NAME)
          workSheets = Program_Choice(workBook)
          Program_Execution(arduino, workSheets)
          continue
        case _ :
          print("Choix invalide.")
          continue
        
  except KeyboardInterrupt:
    print("\nDéconnexion en cours...")
  finally:
    arduino.close()  # Fermer la connexion série
    print("Déconnecté.")


if __name__ == "__main__":
  main()

#while True:
#  #keyboard = Controller()
#  key = input()
#
#  if (key == 32) :
#    while (key == 32):
#      arduino.write((1)).encode()
#      key = input()
#    arduino.write((0)).encode()
#  else :
#    arduino.write((0)).encode()
#
#  #if keyboard.press(Key.space) :
#  #  while ( not keyboard.release(Key.space)) :
#  #    arduino.write((1 + "\n").encode())
#  #  arduino.write((0 + "\n").encode())